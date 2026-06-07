"""Excel KPI Report Exporter - Multi-Sheet Version.

Generates Excel workbook with:
- First sheet "Key KPIs": 5 critical KPIs (even if no data)
  * Fill Rate %
  * OTIF %
  * Days of Supply (AX)
  * Lines Picked per Labor-Hour
  * Error Rate %
  
- Additional sheets: One per collection showing all its KPIs
  * Inbound KPIs
  * Outbound KPIs
  * Inventory KPIs
  * Warehouse Productivity KPIs
  * Employee Productivity KPIs

Each sheet has columns: KPI | Current | Target | Delta | Status | Comment

Output structure:
  output/
    <timestamp>/
      kpi_report.xlsx      - Multi-sheet Excel workbook
      kpi_report.json      - JSON data with header and summary cards
      kpi_onepager.html    - HTML leadership summary report
      
Example: output/20260413_180926/kpi_report.xlsx
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)

# ── Output path ────────────────────────────────────────────────────────────────
_REPO_ROOT  = Path(__file__).resolve().parents[2]   # warehouse_kpi_agent/
_OUTPUT_BASE_DIR = _REPO_ROOT.parent / "output"     # ../output/ (base folder for all exports)

# ── Key KPIs (First Sheet) ───────────────────────────────────────────────────────────
_KEY_KPIS = [
    "fill_rate",
    "otif",
    "days_of_supply",
    "lines_per_labor_hour",
    "error_rate"
]

# ── Collection Mapping ────────────────────────────────────────────────────────────────
_COLLECTION_NAMES = {
    "inbound_parts": "Inbound KPIs",
    "outbound_parts": "Outbound KPIs",
    "inventory_snapshot": "Inventory KPIs",
    "warehouse_productivity": "Warehouse Productivity KPIs",
    "employee_productivity": "Employee Productivity KPIs"
}

# ── Fixed Target Values (Industry Benchmarks) ──────────────────────────────────────
_KPI_TARGETS: dict[str, float] = {
    "fill_rate": 95.0,
    "otif": 92.0,
    "backorder_rate": 5.0,
    "avg_inbound_lead_time": 5.0,
    "on_time_receipts_pct": 95.0,
    "qty_discrepancy_pct": 2.0,
    "days_of_supply": 25.0,
    "stockout_pct": 2.0,
    "lines_per_labor_hour": 6.0,
    "orders_per_day": 50.0,
    "sla_adherence": 95.0,
    "picks_per_hour": 65.0,
    "error_rate": 0.8,
    "overtime_pct": 10.0,
}

# ── KPI Status Rules ───────────────────────────────────────────────────────────
# (green_threshold, amber_threshold, lower_is_better)
_STATUS_RULES: dict[str, tuple[float, float, bool]] = {
    "fill_rate": (90.0, 80.0, False),
    "otif": (90.0, 75.0, False),
    "backorder_rate": (5.0, 10.0, True),
    "avg_inbound_lead_time": (6.0, 8.0, True),
    "on_time_receipts_pct": (90.0, 75.0, False),
    "qty_discrepancy_pct": (3.0, 5.0, True),
    "days_of_supply": (20.0, 15.0, False),
    "stockout_pct": (3.0, 5.0, True),
    "lines_per_labor_hour": (5.5, 4.0, False),
    "orders_per_day": (40.0, 30.0, False),
    "sla_adherence": (90.0, 80.0, False),
    "picks_per_hour": (60.0, 50.0, False),
    "error_rate": (1.0, 2.0, True),
    "overtime_pct": (12.0, 15.0, True),
}

# ── KPI Display Names ──────────────────────────────────────────────────────────
_KPI_NAMES: dict[str, str] = {
    "fill_rate": "Fill Rate %",
    "otif": "OTIF %",
    "backorder_rate": "Backorder Rate %",
    "avg_inbound_lead_time": "Avg Inbound Lead Time (days)",
    "on_time_receipts_pct": "On-Time Receipts %",
    "qty_discrepancy_pct": "Qty Discrepancy %",
    "days_of_supply": "Days of Supply (AX)",
    "stockout_pct": "Stockout %",
    "lines_per_labor_hour": "Lines Picked per Labor-Hour",
    "orders_per_day": "Orders per Day",
    "sla_adherence": "SLA Adherence %",
    "picks_per_hour": "Picks per Hour",
    "error_rate": "Error Rate %",
    "overtime_pct": "Overtime %",
}


def _get_status(kpi_key: str, value: float) -> str:
    """Return GREEN / AMBER / RED based on KPI value and thresholds."""
    if value is None or not isinstance(value, (int, float)):
        return "N/A"
    rule = _STATUS_RULES.get(kpi_key)
    if not rule:
        return "—"
    green_thresh, amber_thresh, lower_better = rule
    if lower_better:
        if value <= green_thresh:
            return "GREEN"
        if value <= amber_thresh:
            return "AMBER"
        return "RED"
    else:
        if value >= green_thresh:
            return "GREEN"
        if value >= amber_thresh:
            return "AMBER"
        return "RED"


def _get_comment(kpi_key: str, current: float, target: float, status: str) -> str:
    """Generate actionable comment based on KPI performance."""
    if current is None:
        return "No data available"
    
    delta = abs(target - current)
    
    # Default comments for status
    if status == "GREEN":
        return "Meeting target"
    elif status == "AMBER":
        return f"Review performance"
    elif status == "RED":
        return f"Immediate action needed"
    
    return "—"


def generate_kpi_excel(
    db_results: list[dict[str, Any]],
    entities: Optional[dict[str, Any]] = None,
    query: str = "",
) -> tuple[Path, Path, Path]:
    """Generate multi-sheet KPI Excel report, JSON, and HTML one-pager.
    
    First sheet: "Key KPIs" - 5 critical KPIs (shows even if no data)
    Additional sheets: One per collection showing all its KPIs

    Args:
        db_results: List of KPI result dicts from registered_kpi_handler.
        entities: Extracted entities {warehouse_id, start_date, end_date}.
        query: Original user query string.

    Returns:
        Tuple of (excel_path, json_path, html_path) - Paths to all generated files.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc

    entities = entities or {}
    warehouse_id = entities.get("warehouse_id") or "All Warehouses"
    start_date = entities.get("start_date") or "—"
    end_date = entities.get("end_date") or "—"
    ts_filename = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Create timestamped output directory for this report
    output_dir = _OUTPUT_BASE_DIR / ts_filename
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Define file paths within the timestamped folder
    excel_path = output_dir / "kpi_report.xlsx"
    json_path = output_dir / "kpi_report.json"
    html_path = output_dir / "kpi_onepager.html"

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Style helpers
    def fill(color: str) -> PatternFill:
        return PatternFill("solid", fgColor=color)

    def font_style(bold=False, color="000000", size=10) -> Font:
        return Font(name="Calibri", bold=bold, color=color, size=size)

    def border_style() -> Border:
        thin = Side(style="thin", color="CCCCCC")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def center_align() -> Alignment:
        return Alignment(horizontal="center", vertical="center")

    def left_align(wrap=False) -> Alignment:
        return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

    # Create KPI lookup dict from results
    kpi_lookup = {r.get("kpi"): r for r in db_results if not isinstance(r.get("value"), list)}

    def _create_kpi_sheet(sheet_name: str, kpi_keys: list[str], title_emoji: str = "📊"):
        """Helper to create a KPI sheet with consistent formatting."""
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = False

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = f"{title_emoji} {sheet_name}"
        ws["A1"].font = font_style(bold=True, color="FFFFFF", size=14)
        ws["A1"].fill = fill("1F4E79")
        ws["A1"].alignment = center_align()
        ws.row_dimensions[1].height = 25

        # Subtitle
        ws.merge_cells("A2:F2")
        ws["A2"] = f"{warehouse_id}  |  {start_date} → {end_date}"
        ws["A2"].font = font_style(size=9, color="666666")
        ws["A2"].alignment = center_align()
        ws.row_dimensions[2].height = 16

        # Spacer
        ws.row_dimensions[3].height = 8

        # Header row
        headers = ["KPI", "Current", "Target", "Delta", "Status", "Comment"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = font_style(bold=True, color="FFFFFF", size=10)
            cell.fill = fill("4472C4")
            cell.alignment = center_align()
            cell.border = border_style()
        ws.row_dimensions[4].height = 20

        # Data rows
        row = 5
        for kpi_key in kpi_keys:
            result = kpi_lookup.get(kpi_key)
            kpi_name = _KPI_NAMES.get(kpi_key, kpi_key)
            target = _KPI_TARGETS.get(kpi_key)

            # Get current value (0 if no data)
            if result and result.get("value") is not None:
                current_value = result.get("value")
            else:
                current_value = None  # Will show as empty

            # Skip if no target defined
            if target is None:
                target = 0  # Default target

            # Calculate delta and status
            if current_value is not None:
                delta = current_value - target
                status = _get_status(kpi_key, current_value)
                
                # For RED status, show absolute value (no negative sign)
                if status == "RED" and delta < 0:
                    delta = abs(delta)
                
                comment = _get_comment(kpi_key, current_value, target, status)
            else:
                delta = None
                status = "—"
                comment = "No data available"

            # Status colors
            status_colors = {
                "GREEN": ("C6EFCE", "006100"),
                "AMBER": ("FFEB9C", "9C6500"),
                "RED": ("FFC7CE", "9C0006"),
                "—": ("FFFFFF", "666666"),
            }
            status_fill, status_font = status_colors.get(status, ("FFFFFF", "000000"))

            # Write data
            data = [kpi_name, current_value, target, delta, status, comment]

            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=value if value is not None else "—")
                cell.border = border_style()

                # Format current, target, delta as numbers
                if col in [2, 3, 4] and isinstance(value, (int, float)):
                    cell.number_format = "0.0"
                    cell.alignment = center_align()
                elif col == 5:  # Status column
                    cell.fill = fill(status_fill)
                    cell.font = font_style(bold=True, color=status_font, size=10)
                    cell.alignment = center_align()
                elif col == 6:  # Comment column
                    cell.alignment = left_align(wrap=True)
                    cell.font = font_style(size=9)
                else:
                    cell.alignment = left_align()
                    cell.font = font_style(size=10)

            ws.row_dimensions[row].height = 22
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 32  # KPI
        ws.column_dimensions['B'].width = 12  # Current
        ws.column_dimensions['C'].width = 12  # Target
        ws.column_dimensions['D'].width = 12  # Delta
        ws.column_dimensions['E'].width = 12  # Status
        ws.column_dimensions['F'].width = 50  # Comment

    # ──────────────────────────────────────────────────────────────────────────
    # Sheet 1: Key KPIs (always include these 5, even if no data)
    # ──────────────────────────────────────────────────────────────────────────
    _create_kpi_sheet("Key KPIs", _KEY_KPIS, "⭐")

    # ──────────────────────────────────────────────────────────────────────────
    # Additional sheets: One per collection
    # ──────────────────────────────────────────────────────────────────────────
    # Group KPIs by collection
    collection_kpis = {}
    for result in db_results:
        if isinstance(result.get("value"), list):  # Skip list-type KPIs
            continue
        collection = result.get("collection", "unknown")
        kpi_key = result.get("kpi")
        if kpi_key:
            if collection not in collection_kpis:
                collection_kpis[collection] = []
            if kpi_key not in collection_kpis[collection]:
                collection_kpis[collection].append(kpi_key)

    # Create sheet for each collection
    for collection, kpi_keys in sorted(collection_kpis.items()):
        sheet_name = _COLLECTION_NAMES.get(collection, collection.replace("_", " ").title())
        _create_kpi_sheet(sheet_name, kpi_keys, "📊")

    # Save Excel
    wb.save(excel_path)
    logger.info("KPI Excel saved: %s", excel_path)
    
    # ──────────────────────────────────────────────────────────────────────────
    # Generate JSON output alongside Excel
    # ──────────────────────────────────────────────────────────────────────────
    json_result = None
    try:
        json_result = _generate_kpi_json(
            db_results=db_results,
            entities=entities,
            output_path=json_path,
            kpi_lookup=kpi_lookup
        )
        logger.info("KPI JSON saved: %s", json_result)
    except Exception as exc:
        logger.error("JSON export failed: %s", exc)
    
    # ──────────────────────────────────────────────────────────────────────────
    # Generate HTML One-Pager
    # ──────────────────────────────────────────────────────────────────────────
    html_result = None
    try:
        html_result = _generate_html_onepager(
            db_results=db_results,
            entities=entities,
            output_path=html_path,
            kpi_lookup=kpi_lookup
        )
        logger.info("KPI HTML One-Pager saved: %s", html_result)
    except Exception as exc:
        logger.error("HTML one-pager export failed: %s", exc)
    
    return excel_path, json_result, html_result


def _generate_kpi_json(
    db_results: list[dict[str, Any]],
    entities: dict[str, Any],
    output_path: Path,
    kpi_lookup: dict[str, Any]
) -> Path:
    """Generate JSON output matching the specified schema.
    
    Output format:
    {
      "header": {
        "period": "2025-06-01 → 2025-06-30",
        "warehouses": ["WH-01"],
        "status": "complete"
      },
      "summary_cards": [
        {
          "name": "Fill Rate %",
          "current": 77.58,
          "target": 95.0,
          "delta": -17.42,
          "status": "RED",
          "comment": "Below target - review stockouts"
        }
      ]
    }
    """
    warehouse_id = entities.get("warehouse_id") or "All Warehouses"
    start_date = entities.get("start_date") or "—"
    end_date = entities.get("end_date") or "—"
    
    # Build header
    header = {
        "period": f"{start_date} → {end_date}",
        "warehouses": [warehouse_id] if warehouse_id != "All Warehouses" else [],
        "status": "complete"
    }
    
    # Build summary_cards for Key KPIs only
    summary_cards = []
    
    for kpi_key in _KEY_KPIS:
        result = kpi_lookup.get(kpi_key)
        kpi_name = _KPI_NAMES.get(kpi_key, kpi_key)
        target = _KPI_TARGETS.get(kpi_key)
        
        # Get current value
        if result and result.get("value") is not None:
            current_value = result.get("value")
        else:
            current_value = None
        
        # Calculate delta and status
        if current_value is not None and target is not None:
            delta = current_value - target
            status = _get_status(kpi_key, current_value)
            
            # For RED status, show absolute value (no negative sign)
            if status == "RED" and delta < 0:
                delta = abs(delta)
            
            # Round delta to 2 decimal places for cleaner JSON
            delta = round(delta, 2) if delta is not None else None
            
            comment = _get_comment(kpi_key, current_value, target, status)
        else:
            delta = None
            status = "—"
            comment = "No data available"
        
        # Build card
        card = {
            "name": kpi_name,
            "current": current_value if current_value is not None else "—",
            "target": target if target is not None else 0,
            "delta": delta,
            "status": status,
            "comment": comment
        }
        
        summary_cards.append(card)
    
    # Build final JSON structure
    output_data = {
        "header": header,
        "summary_cards": summary_cards
    }
    
    # Write to file with proper formatting
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)
    
    return output_path


def _generate_html_onepager(
    db_results: list[dict[str, Any]],
    entities: dict[str, Any],
    output_path: Path,
    kpi_lookup: dict[str, Any]
) -> Path:
    """Generate HTML One-Page Leadership Summary.
    
    Includes:
    - Header with period, warehouses, status
    - Top Summary KPIs (dashboard cards)
    - Collection sections (tiles)
    - Tabular KPI Summary
    - Insights & Recommended Actions
    """
    warehouse_id = entities.get("warehouse_id") or "All Warehouses"
    start_date = entities.get("start_date") or "—"
    end_date = entities.get("end_date") or "—"
    
    # Determine overall status
    overall_status = "GREEN"
    red_count = sum(1 for kpi_key in _KEY_KPIS 
                    if kpi_lookup.get(kpi_key) and 
                    _get_status(kpi_key, kpi_lookup[kpi_key].get("value", 0)) == "RED")
    amber_count = sum(1 for kpi_key in _KEY_KPIS 
                      if kpi_lookup.get(kpi_key) and 
                      _get_status(kpi_key, kpi_lookup[kpi_key].get("value", 0)) == "AMBER")
    
    if red_count > 0:
        overall_status = "RED"
    elif amber_count > 0:
        overall_status = "AMBER"
    
    # Group KPIs by collection for sections
    collection_kpis_data = {}
    for result in db_results:
        if isinstance(result.get("value"), list):
            continue
        collection = result.get("collection", "unknown")
        if collection not in collection_kpis_data:
            collection_kpis_data[collection] = []
        collection_kpis_data[collection].append(result)
    
    # Build HTML
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Warehouse KPI Leadership Summary</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ 
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
            background: #f5f5f5; 
            padding: 20px;
            font-size: 12px;
        }}
       .container {{ max-width: 1400px; margin: 0 auto; background: white; padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
        
        /* Header */
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        .header h1 {{ font-size: 24px; margin-bottom: 5px; }}
        .header-info {{ font-size: 13px; }}
        .status-badge {{
            padding: 8px 16px;
            border-radius: 20px;
            font-weight: bold;
            font-size: 14px;
        }}
        .status-GREEN {{ background: #10b981; color: white; }}
        .status-AMBER {{ background: #f59e0b; color: white; }}
        .status-RED {{ background: #ef4444; color: white; }}
        
        /* Top Summary Cards */
        .top-summary {{
            display: grid;
            grid-template-columns: repeat(5, 1fr);
            gap: 15px;
            margin-bottom: 20px;
        }}
        .summary-card {{
            background: white;
            border: 1px solid #e5e7eb;
            border-radius: 8px;
            padding: 15px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }}
        .summary-card h3 {{ font-size: 11px; color: #6b7280; margin-bottom: 8px; text-transform: uppercase; }}
        .summary-card  .value {{ font-size: 28px; font-weight: bold; margin-bottom: 5px; }}
        .summary-card .target {{ font-size: 11px; color: #9ca3af; }}
        .summary-card .delta {{ font-size: 12px; font-weight: 600; margin-top: 5px; }}
        .delta-positive {{ color: #10b981; }}
        .delta-negative {{ color: #ef4444; }}
        
        /* Sections */
        .sections {{
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 15px;
            margin-bottom: 20px;
        }}
        .section-tile {{
            background: white;
            border: 1px solid #e5e7eb;
            border-radius: 8px;
            padding: 15px;
        }}
        .section-tile h2 {{
            font-size: 16px;
            color: #111827;
            margin-bottom: 12px;
            padding-bottom: 8px;
            border-bottom: 2px solid #e5e7eb;
        }}
        .section-tile .kpi-row {{
            display: flex;
            justify-content: space-between;
            padding: 6px 0;
            border-bottom: 1px solid #f3f4f6;
        }}
        .section-tile .kpi-row:last-child {{ border-bottom: none; }}
        .section-tile .kpi-name {{ color: #4b5563; font-size: 11px; }}
        .section-tile .kpi-value {{ font-weight: 600; font-size: 12px; }}
        
        /* KPI Table */
        .kpi-table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }}
        .kpi-table th {{
            background: #f9fafb;
            padding: 12px;
            text-align: left;
            font-size: 11px;
            font-weight: 600;
            color: #374151;
            border-bottom: 2px solid #e5e7eb;
        }}
        .kpi-table td {{
            padding: 10px 12px;
            border-bottom: 1px solid #f3f4f6;
            font-size: 11px;
        }}
        .kpi-table tr:last-child td {{ border-bottom: none; }}
        .kpi-table tr:hover {{ background: #f9fafb; }}
        
        /* Insights */
        .insights {{
            background: #fff7ed;
            border-left: 4px solid #f59e0b;
            padding: 15px;
            border-radius: 4px;
        }}
        .insights h2 {{
            font-size: 16px;
            color: #92400e;
            margin-bottom: 10px;
        }}
        .insights ul {{
            list-style: none;
            padding-left: 0;
        }}
        .insights li {{
            padding: 6px 0;
            color: #78350f;
            font-size: 11px;
            line-height: 1.5;
        }}
        .insights li:before {{
            content: "▸ ";
            color: #f59e0b;
            font-weight: bold;
            margin-right: 8px;
        }}
        
        .status-dot {{
            display: inline-block;
            width: 8px;
            height: 8px;
            border-radius: 50%;
            margin-right: 6px;
        }}
        .status-dot.green {{ background: #10b981; }}
        .status-dot.amber {{ background: #f59e0b; }}
        .status-dot.red {{ background: #ef4444; }}
    </style>
</head>
<body>
    <div class="container">
        <!-- Header -->
        <div class="header">
            <div>
                <h1>🏭 Warehouse KPI Leadership Summary</h1>
                <div class="header-info">
                    <strong>Period:</strong> {start_date} → {end_date} &nbsp;|&nbsp;
                    <strong>Warehouse:</strong> {warehouse_id}
                </div>
            </div>
            <div class="status-badge status-{overall_status}">{overall_status}</div>
        </div>
        
        <!-- Top Summary Cards -->
        <div class="top-summary">
"""
    
    # Add top summary cards for Key KPIs
    for kpi_key in _KEY_KPIS:
        result = kpi_lookup.get(kpi_key)
        kpi_name = _KPI_NAMES.get(kpi_key, kpi_key)
        target = _KPI_TARGETS.get(kpi_key, 0)
        
        if result and result.get("value") is not None:
            current = result.get("value")
            delta = current - target
            status = _get_status(kpi_key, current)
            delta_class = "delta-positive" if delta >= 0 else "delta-negative"
            delta_sign = "+" if delta >= 0 else ""
            
            status_dot_class = "green" if status == "GREEN" else ("amber" if status == "AMBER" else "red")
            delta_display = f"{delta:.1f}" if isinstance(delta, (int, float)) else str(delta)
        else:
            current = "—"
            delta = 0
            delta_display = "—"
            delta_class = ""
            delta_sign = ""
            status_dot_class = "amber"
        
        html_content += f"""            <div class="summary-card">
                <h3>{kpi_name}</h3>
                <div class="value"><span class="status-dot {status_dot_class}"></span>{current}</div>
                <div class="target">Target: {target}</div>
                <div class="delta {delta_class}">{delta_sign}{delta_display}</div>
            </div>
"""
    
    html_content += """        </div>
        
        <!-- Collection Sections -->
        <div class="sections">
"""
    
    # Add collection tiles
    collection_names_display = {
        "inbound_parts": "📦 Inbound",
        "outbound_parts": "🚚 Outbound",
        "inventory_snapshot": "📊 Inventory",
        "warehouse_productivity": "🏭 Warehouse Productivity",
        "employee_productivity": "👷 Employee Productivity"
    }
    
    for collection, display_name in collection_names_display.items():
        if collection in collection_kpis_data:
            html_content += f"""            <div class="section-tile">
                <h2>{display_name}</h2>
"""
            for result in collection_kpis_data[collection][:5]:  # Show top 5 KPIs per section
                kpi_name = result.get("name", result.get("kpi", ""))
                value = result.get("value")
                unit = result.get("unit", "")
                
                if isinstance(value, (int, float)):
                    value_str = f"{value:.1f} {unit}" if unit != "%" else f"{value:.1f}%"
                else:
                    value_str = "—"
                
                html_content += f"""                <div class="kpi-row">
                    <span class="kpi-name">{kpi_name}</span>
                    <span class="kpi-value">{value_str}</span>
                </div>
"""
            html_content += """            </div>
"""
    
    html_content += """        </div>
        
        <!-- KPI Table -->
        <h2 style="margin-bottom: 10px; color: #111827;">Tabular KPI Summary</h2>
        <table class="kpi-table">
            <thead>
                <tr>
                    <th>KPI</th>
                    <th>Current</th>
                    <th>Target</th>
                    <th>Delta</th>
                    <th>Status</th>
                    <th>Comment</th>
                </tr>
            </thead>
            <tbody>
"""
    
    # Add table rows for Key KPIs
    for kpi_key in _KEY_KPIS:
        result = kpi_lookup.get(kpi_key)
        kpi_name = _KPI_NAMES.get(kpi_key, kpi_key)
        target = _KPI_TARGETS.get(kpi_key, 0)
        
        if result and result.get("value") is not None:
            current = result.get("value")
            delta = current - target
            status = _get_status(kpi_key, current)
            comment = _get_comment(kpi_key, current, target, status)
            
            if status == "RED" and delta < 0:
                delta = abs(delta)
            
            status_dot_class = "green" if status == "GREEN" else ("amber" if status == "AMBER" else "red")
            delta_sign = "+" if delta >= 0 else ""
            current_display = f"{current:.1f}" if isinstance(current, (int, float)) else str(current)
            target_display = f"{target:.1f}" if isinstance(target, (int, float)) else str(target)
            delta_display = f"{delta:.1f}" if isinstance(delta, (int, float)) else str(delta)
        else:
            current = "—"
            current_display = "—"
            delta = "—"
            delta_display = "—"
            status = "—"
            comment = "No data available"
            status_dot_class = "amber"
            delta_sign = ""
            target_display = f"{target:.1f}" if isinstance(target, (int, float)) else str(target)
        
        html_content += f"""                <tr>
                    <td>{kpi_name}</td>
                    <td>{current_display}</td>
                    <td>{target_display}</td>
                    <td>{delta_sign}{delta_display}</td>
                    <td><span class="status-dot {status_dot_class}"></span>{status}</td>
                    <td>{comment}</td>
                </tr>
"""
    
    html_content += """            </tbody>
        </table>
        
        <!-- Insights & Actions -->
        <div class="insights">
            <h2>💡 Insights & Recommended Actions</h2>
            <ul>
"""
    
    # Generate insights based on KPI status
    insights = []
    for kpi_key in _KEY_KPIS:
        result = kpi_lookup.get(kpi_key)
        if result and result.get("value") is not None:
            current = result.get("value")
            target = _KPI_TARGETS.get(kpi_key, 0)
            status = _get_status(kpi_key, current)
            kpi_name = _KPI_NAMES.get(kpi_key, kpi_key)
            
            if status == "RED":
                insights.append(f"<strong>{kpi_name}:</strong> Critical - current at {current:.1f}, target {target:.1f}. Immediate action required.")
            elif status == "AMBER":
                insights.append(f"<strong>{kpi_name}:</strong> Below target ({current:.1f} vs {target:.1f}). Review processes and implement improvements.")
    
    if not insights:
        insights.append("All key metrics are within target range. Continue monitoring performance.")
    
    for insight in insights[:5]:  # Show top 5 insights
        html_content += f"""                <li>{insight}</li>
"""
    
    html_content += """            </ul>
        </div>
    </div>
</body>
</html>
"""
    
    # Write HTML file
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    return output_path
